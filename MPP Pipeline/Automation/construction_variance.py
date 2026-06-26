"""
Stage E — Construction Variance Analysis  (Methodology A)
=========================================================
Rolls up the Construction portion of the schedule by task name within each
reporting bucket, computes span-based duration variance, and produces the
ranking tables used in the brief.

Span definition (Methodology A, step 2):
  Span = elapsed working time from earliest start to latest finish across a
  rolled-up line's instances, on the project's Mon-Fri 8-hour calendar.
  For tasks not yet complete, the actual span runs to the current scheduled
  (forecast) finish.

Faithful implementation detail (validated against the Harrison Top-24):
  - Single-instance lines take the span straight from MS Project's own
    Duration / BaselineDuration field — the exact, possibly-fractional
    working-time MS Project computed (e.g. Detention Pond 199.1 d).
  - Multi-instance lines take the combined-window span via a Mon-Fri
    business-day count between the earliest start and latest finish
    (integer days; matches Cut/Fill 109 d, Layout Footings 46 d).
  Every fractional Top-24 row is single-instance and every multi-instance
  row is integer, so this reproduces the published figures exactly.

Variance:
  Absolute Variance (d) = Actual Span - Baseline Span
  Percent Variance      = Abs / Baseline, "n/a" when Baseline < 0.5 d
  Net bucket variance    = arithmetic sum of task absolute variances

Inputs:
  - project_config.json
  - Stage C output (output_root/stage_c/snapshots/<snapshot>.parquet)

Outputs (output_root/stage_e/):
  - construction_variance_full.parquet   — every rolled-up line
  - top_n_all_buckets.parquet            — Top-N across all buckets
  - per_building_top.parquet             — per-building Top-N, grouped by phase
  - bucket_summary.parquet               — net variance per bucket
  - variance_report.json
  - Construction_Variance_<snapshot>.xlsx  — J New Town styled workbook

No Java required — reads the Stage C parquet directly.

Usage:
  python construction_variance.py [--config project_config.json] [--snapshot <stem>]
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd


# ---------------------------------------------------------------------------
# Config + snapshot selection
# ---------------------------------------------------------------------------

def load_config(config_path: str) -> dict:
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def find_snapshot(cfg: dict, snapshot_arg: str | None) -> Path:
    output_root = Path(cfg["paths"]["output_root"])
    snap_dir = output_root / "stage_c" / "snapshots"
    if not snap_dir.exists():
        raise FileNotFoundError(
            f"Stage C output not found at {snap_dir}. Run extract_snapshots.py first."
        )
    parquets = sorted(snap_dir.glob("*.parquet"), key=lambda p: p.stem)
    if not parquets:
        raise FileNotFoundError(f"No snapshot parquet files in {snap_dir}.")
    if snapshot_arg:
        match = snap_dir / f"{snapshot_arg}.parquet"
        if not match.exists():
            raise FileNotFoundError(
                f"Snapshot '{snapshot_arg}' not found. Available: {[p.stem for p in parquets]}"
            )
        return match
    return parquets[-1]


# ---------------------------------------------------------------------------
# Working-time helper (Mon-Fri business-day calendar, finish-inclusive)
# ---------------------------------------------------------------------------

def working_days(start, finish, weekmask: str = "1111100", holidays=None):
    """
    Working days between two dates, counting the finish day itself.
    Returns None if either bound is missing, 0.0 if finish precedes start.
    """
    if start is None or finish is None or pd.isna(start) or pd.isna(finish):
        return None
    s = np.datetime64(pd.Timestamp(start).date(), "D")
    f = np.datetime64(pd.Timestamp(finish).date(), "D")
    if f < s:
        return 0.0
    hol = np.array(holidays or [], dtype="datetime64[D]")
    return float(np.busday_count(s, f + np.timedelta64(1, "D"),
                                 weekmask=weekmask, holidays=hol))


# ---------------------------------------------------------------------------
# Resource aggregation
# ---------------------------------------------------------------------------

def merge_resources(series: pd.Series) -> str:
    """Collect distinct resource tokens across a rolled-up line's instances."""
    seen = []
    for val in series.dropna():
        for tok in str(val).split(";"):
            t = tok.strip()
            if t and t not in seen:
                seen.append(t)
    return "; ".join(seen)


# ---------------------------------------------------------------------------
# Core analysis (pure function — fully unit-testable)
# ---------------------------------------------------------------------------

def build_construction_variance(tasks_df: pd.DataFrame, cfg: dict):
    """
    Returns
    -------
    (full_df, top_n_df, per_building_df, bucket_summary_df, report)
    """
    buyout_end = cfg["schedule"]["buyout_task_id_range_end"]
    cv = cfg["construction_variance"]
    buckets = cv["buckets"]
    bucket_lookup = {b.strip().lower(): b for b in buckets}      # normalized -> canonical
    bucket_order = {b: i for i, b in enumerate(buckets)}
    top_n = cv.get("top_n_ranking", 24)
    per_b_n = cv.get("per_building_top_n", 10)
    min_base = cv.get("baseline_span_min_days", 0.5)

    cal = cfg.get("working_calendar", {})
    weekmask = cal.get("weekmask", "1111100")
    holidays = cal.get("holidays", [])

    building_names = set(cfg["buildings"]["names"])
    # building -> phase label
    bld_phase = {}
    for ph in cfg["buildings"].get("phases", []):
        for b in ph.get("buildings", []):
            bld_phase[b] = ph.get("label", f"Phase {ph.get('phase_id','')}")

    # ── Scope to construction ─────────────────────────────────────────────
    df = tasks_df[tasks_df["uid"] > buyout_end].copy()

    # ── uid -> (name, parent_uid) for bucket climbs ───────────────────────
    name_by_uid = {}
    parent_by_uid = {}
    for row in df.itertuples(index=False):
        uid = int(row.uid)
        name_by_uid[uid] = str(row.name) if pd.notna(row.name) else ""
        p = getattr(row, "parent_uid")
        parent_by_uid[uid] = int(p) if pd.notna(p) else None

    def resolve_bucket(leaf_uid: int):
        """First ancestor (incl. leaf) whose name matches a configured bucket."""
        cur = leaf_uid
        visited = set()
        while cur is not None and cur not in visited:
            visited.add(cur)
            nm = name_by_uid.get(cur, "").strip().lower()
            if nm in bucket_lookup:
                return bucket_lookup[nm]
            cur = parent_by_uid.get(cur)
        return None

    # ── Leaf table with effective actual dates and resolved bucket ────────
    leaves = df[~df["is_summary"]].copy()
    leaves["bucket"] = [resolve_bucket(int(u)) for u in leaves["uid"]]
    unbucketed = int(leaves["bucket"].isna().sum())
    leaves = leaves[leaves["bucket"].notna()].copy()

    # Effective actual dates: fall back to scheduled (forecast) when missing
    leaves["eff_actual_start"] = leaves["actual_start"].where(
        leaves["actual_start"].notna(), leaves["sched_start"])
    leaves["eff_actual_finish"] = leaves["actual_finish"].where(
        leaves["actual_finish"].notna(), leaves["sched_finish"])

    # ── Roll up by (bucket, name) ─────────────────────────────────────────
    rows = []
    for (bucket, name), grp in leaves.groupby(["bucket", "name"], sort=False):
        n = len(grp)

        if n == 1:
            r = grp.iloc[0]
            base = r["baseline_duration"]
            base_span = float(base) if pd.notna(base) else \
                (working_days(r["baseline_start"], r["baseline_finish"], weekmask, holidays) or 0.0)
            act = r["duration"]
            act_span = float(act) if pd.notna(act) else \
                (working_days(r["eff_actual_start"], r["eff_actual_finish"], weekmask, holidays) or 0.0)
        else:
            bsd = working_days(grp["baseline_start"].min(), grp["baseline_finish"].max(),
                               weekmask, holidays)
            base_span = bsd if bsd is not None else 0.0
            asd = working_days(grp["eff_actual_start"].min(), grp["eff_actual_finish"].max(),
                               weekmask, holidays)
            act_span = asd if asd is not None else 0.0

        abs_var = act_span - base_span
        pct_var = (abs_var / base_span) if base_span >= min_base else np.nan

        phase = bld_phase.get(bucket, "") if bucket in building_names else ""

        rows.append({
            "bucket":        bucket,
            "phase":         phase,
            "task_name":     name,
            "resources":     merge_resources(grp["resources"]),
            "instances":     n,
            "baseline_span": round(base_span, 1),
            "actual_span":   round(act_span, 1),
            "abs_variance":  round(abs_var, 1),
            "pct_variance":  pct_var,
        })

    full_df = pd.DataFrame(rows)

    # ── Top-N across all buckets ──────────────────────────────────────────
    if not full_df.empty:
        top_n_df = (full_df.sort_values("abs_variance", ascending=False)
                    .head(top_n).reset_index(drop=True))
        top_n_df.insert(0, "rank", range(1, len(top_n_df) + 1))
    else:
        top_n_df = full_df.copy()

    # ── Per-building Top-N, grouped by phase ──────────────────────────────
    pb_frames = []
    if not full_df.empty:
        bld_df = full_df[full_df["bucket"].isin(building_names)]
        for bucket, grp in bld_df.groupby("bucket", sort=False):
            top = grp.sort_values("abs_variance", ascending=False).head(per_b_n).copy()
            top.insert(0, "building_rank", range(1, len(top) + 1))
            pb_frames.append(top)
    if pb_frames:
        per_building_df = pd.concat(pb_frames, ignore_index=True)
        # order by phase, then bucket order, then rank
        per_building_df["_phase_key"] = per_building_df["phase"]
        per_building_df["_bucket_key"] = per_building_df["bucket"].map(bucket_order)
        per_building_df = (per_building_df
                           .sort_values(["_phase_key", "_bucket_key", "building_rank"])
                           .drop(columns=["_phase_key", "_bucket_key"])
                           .reset_index(drop=True))
    else:
        per_building_df = pd.DataFrame()

    # ── Bucket summary (net variance, over/under decomposition) ───────────
    bs_rows = []
    if not full_df.empty:
        for bucket, grp in full_df.groupby("bucket", sort=False):
            over = grp[grp["abs_variance"] > 0]
            under = grp[grp["abs_variance"] < 0]
            bs_rows.append({
                "bucket":        bucket,
                "lines":         len(grp),
                "instances":     int(grp["instances"].sum()),
                "over_count":    len(over),
                "over_days":     round(over["abs_variance"].sum(), 1),
                "under_count":   len(under),
                "under_days":    round(under["abs_variance"].sum(), 1),
                "net_variance":  round(grp["abs_variance"].sum(), 1),
            })
    bucket_summary_df = pd.DataFrame(bs_rows)
    if not bucket_summary_df.empty:
        bucket_summary_df["_k"] = bucket_summary_df["bucket"].map(bucket_order)
        bucket_summary_df = (bucket_summary_df.sort_values("_k")
                             .drop(columns="_k").reset_index(drop=True))

    # ── Report ────────────────────────────────────────────────────────────
    sitework_net = None
    if not bucket_summary_df.empty:
        sw = bucket_summary_df[bucket_summary_df["bucket"].str.lower() == "site work"]
        if not sw.empty:
            r = sw.iloc[0]
            sitework_net = {
                "net": float(r["net_variance"]),
                "over_count": int(r["over_count"]), "over_days": float(r["over_days"]),
                "under_count": int(r["under_count"]), "under_days": float(r["under_days"]),
            }

    report = {
        "construction_task_count": len(df),
        "construction_leaf_count": int((~df["is_summary"]).sum()),
        "bucketed_leaf_count":     len(leaves),
        "unbucketed_leaf_count":   unbucketed,
        "rolled_up_line_count":    len(full_df),
        "buckets_with_data":       int(full_df["bucket"].nunique()) if not full_df.empty else 0,
        "site_work_net":           sitework_net,
    }

    return full_df, top_n_df, per_building_df, bucket_summary_df, report


# ---------------------------------------------------------------------------
# Styled workbook  (J New Town template)
# ---------------------------------------------------------------------------

NAVY = "FF1F4E78"
ZEBRA = "FFF2F6FB"
REDTINT = "FFFCE4E4"
GRAY = "FF606060"
WHITE = "FFFFFFFF"


def _style_variance_sheet(ws, title, subtitle, df, text_cols):
    """
    Write a variance table with J New Town styling.

    text_cols : list of (header, df_key) for the leading text/rank columns.
    The trailing columns are always:
        Instances, Baseline Span (d), Actual Span (d), Abs Var (d), % Var
    where Abs Var and % Var are written as Excel formulas.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    navy_fill = PatternFill("solid", fgColor=NAVY)
    zebra_fill = PatternFill("solid", fgColor=ZEBRA)
    red_fill = PatternFill("solid", fgColor=REDTINT)
    hdr_font = Font(name="Calibri", bold=True, color=WHITE)
    title_font = Font(name="Calibri", size=14, bold=True, color=NAVY)
    sub_font = Font(name="Calibri", size=10, color=GRAY)
    base_font = Font(name="Calibri")
    bold_font = Font(name="Calibri", bold=True)

    # Title + subtitle
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A2"] = subtitle
    ws["A2"].font = sub_font

    header_row = 4
    trailing = ["Instances", "Baseline Span (d)", "Actual Span (d)",
                "Abs Var (d)", "% Var"]
    headers = [h for h, _ in text_cols] + trailing
    n_text = len(text_cols)
    n_cols = len(headers)

    # Column letters for formula references
    col_instances = n_text + 1
    col_baseline = n_text + 2
    col_actual = n_text + 3
    col_absvar = n_text + 4
    col_pct = n_text + 5
    L_base = get_column_letter(col_baseline)
    L_act = get_column_letter(col_actual)
    L_abs = get_column_letter(col_absvar)

    # Header
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = hdr_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data
    for i, (_, rec) in enumerate(df.iterrows()):
        r = header_row + 1 + i
        striped = (i % 2 == 1)
        fill = zebra_fill if striped else None

        # text/rank cols
        for c, (_, key) in enumerate(text_cols, start=1):
            val = rec.get(key, "")
            if pd.isna(val):
                val = ""
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = base_font
            # rank-like first column centered if header is Rank/#
            hdr = text_cols[c - 1][0].lower()
            cell.alignment = Alignment(
                horizontal="center" if hdr in ("rank", "#") else "left")
            if fill:
                cell.fill = fill

        # Instances
        cell = ws.cell(row=r, column=col_instances, value=int(rec["instances"]))
        cell.font = base_font
        cell.number_format = "0"
        cell.alignment = Alignment(horizontal="center")
        if fill:
            cell.fill = fill

        # Baseline span
        cell = ws.cell(row=r, column=col_baseline, value=float(rec["baseline_span"]))
        cell.font = base_font
        cell.number_format = "0.0"
        cell.alignment = Alignment(horizontal="right")
        if fill:
            cell.fill = fill

        # Actual span
        cell = ws.cell(row=r, column=col_actual, value=float(rec["actual_span"]))
        cell.font = base_font
        cell.number_format = "0.0"
        cell.alignment = Alignment(horizontal="right")
        if fill:
            cell.fill = fill

        # Abs Var (formula, bold, red-tint always)
        cell = ws.cell(row=r, column=col_absvar, value=f"={L_act}{r}-{L_base}{r}")
        cell.font = bold_font
        cell.number_format = "0.0"
        cell.alignment = Alignment(horizontal="right")
        cell.fill = red_fill

        # % Var (formula, suppress to n/a under min baseline)
        cell = ws.cell(row=r, column=col_pct,
                       value=f'=IF({L_base}{r}<0.5,"n/a",{L_abs}{r}/{L_base}{r})')
        cell.font = base_font
        cell.number_format = "0.0%"
        cell.alignment = Alignment(horizontal="right")
        if fill:
            cell.fill = fill

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    _autosize(ws, headers, n_cols)


def _style_bucket_summary(ws, title, subtitle, df):
    from openpyxl.styles import Font, PatternFill, Alignment

    navy_fill = PatternFill("solid", fgColor=NAVY)
    zebra_fill = PatternFill("solid", fgColor=ZEBRA)
    red_fill = PatternFill("solid", fgColor=REDTINT)
    hdr_font = Font(name="Calibri", bold=True, color=WHITE)
    title_font = Font(name="Calibri", size=14, bold=True, color=NAVY)
    sub_font = Font(name="Calibri", size=10, color=GRAY)
    base_font = Font(name="Calibri")
    bold_font = Font(name="Calibri", bold=True)

    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A2"] = subtitle
    ws["A2"].font = sub_font

    headers = ["Bucket", "Lines", "Instances", "Over (n)", "Over (d)",
               "Under (n)", "Under (d)", "Net Variance (d)"]
    hr = 4
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.font = hdr_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, (_, rec) in enumerate(df.iterrows()):
        r = hr + 1 + i
        fill = zebra_fill if i % 2 == 1 else None
        vals = [rec["bucket"], int(rec["lines"]), int(rec["instances"]),
                int(rec["over_count"]), float(rec["over_days"]),
                int(rec["under_count"]), float(rec["under_days"]),
                float(rec["net_variance"])]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            if c == 1:
                cell.font = base_font
                cell.alignment = Alignment(horizontal="left")
            elif c == 8:
                cell.font = bold_font
                cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="right")
                cell.fill = red_fill
                if fill:
                    pass  # red-tint wins on net column
                continue
            else:
                cell.font = base_font
                cell.number_format = "0.0" if c in (5, 7) else "0"
                cell.alignment = Alignment(horizontal="right")
            if fill:
                cell.fill = fill

    ws.freeze_panes = ws.cell(row=hr + 1, column=1)
    _autosize(ws, headers, len(headers))


def _autosize(ws, headers, n_cols):
    from openpyxl.utils import get_column_letter
    for c in range(1, n_cols + 1):
        letter = get_column_letter(c)
        maxlen = len(str(headers[c - 1]))
        for row in ws.iter_rows(min_col=c, max_col=c, min_row=5):
            for cell in row:
                if cell.value is not None and not str(cell.value).startswith("="):
                    maxlen = max(maxlen, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(maxlen + 2, 8), 40)


def write_workbook(path, cfg, snapshot, full_df, top_n_df, per_building_df, bucket_summary_df):
    from openpyxl import Workbook

    status = cfg["project"].get("analysis_status_date", "")
    subtitle = (f"Methodology A — span-based. Variance = Actual − Baseline span "
                f"(working days). Status {status}.")

    wb = Workbook()

    # Top-N
    ws = wb.active
    ws.title = f"Top {len(top_n_df)}"
    _style_variance_sheet(
        ws, f"Top {len(top_n_df)} Duration Variances — Across All Buckets",
        subtitle, top_n_df,
        text_cols=[("Rank", "rank"), ("Bucket", "bucket"),
                   ("Task Name", "task_name"), ("Resources", "resources")])

    # Bucket summary
    ws = wb.create_sheet("Bucket Summary")
    _style_bucket_summary(ws, "Construction Variance — Bucket Summary",
                          subtitle, bucket_summary_df)

    # Per-building
    if not per_building_df.empty:
        ws = wb.create_sheet("Per-Building Top")
        _style_variance_sheet(
            ws, "Per-Building Top Variances (grouped by phase)",
            subtitle, per_building_df,
            text_cols=[("#", "building_rank"), ("Phase", "phase"),
                       ("Building", "bucket"), ("Task Name", "task_name"),
                       ("Resources", "resources")])

    # Full detail
    ws = wb.create_sheet("Full Detail")
    full_sorted = full_df.sort_values(["bucket", "abs_variance"],
                                      ascending=[True, False]).reset_index(drop=True)
    _style_variance_sheet(
        ws, "Construction Variance — Full Detail (all rolled-up lines)",
        subtitle, full_sorted,
        text_cols=[("Bucket", "bucket"), ("Phase", "phase"),
                   ("Task Name", "task_name"), ("Resources", "resources")])

    wb.save(path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Stage E — construction variance")
    parser.add_argument("--config", default="project_config.json")
    parser.add_argument("--snapshot", default=None)
    args = parser.parse_args()

    cfg = load_config(args.config)
    project_name = cfg["project"]["name"]
    output_root = Path(cfg["paths"]["output_root"])
    stage_dir = output_root / "stage_e"
    stage_dir.mkdir(parents=True, exist_ok=True)

    snapshot_path = find_snapshot(cfg, args.snapshot)

    print(f"\n{'='*60}")
    print(f"  Stage E — Construction Variance Analysis")
    print(f"  Project  : {project_name}")
    print(f"  Snapshot : {snapshot_path.stem}")
    print(f"{'='*60}\n")

    tasks_df = pd.read_parquet(snapshot_path)
    full_df, top_n_df, per_building_df, bucket_summary_df, report = \
        build_construction_variance(tasks_df, cfg)

    # Persist parquet
    full_df.to_parquet(stage_dir / "construction_variance_full.parquet", index=False)
    top_n_df.to_parquet(stage_dir / "top_n_all_buckets.parquet", index=False)
    if not per_building_df.empty:
        per_building_df.to_parquet(stage_dir / "per_building_top.parquet", index=False)
    bucket_summary_df.to_parquet(stage_dir / "bucket_summary.parquet", index=False)

    # Styled workbook
    xlsx_path = stage_dir / f"Construction_Variance_{snapshot_path.stem}.xlsx"
    write_workbook(xlsx_path, cfg, snapshot_path.stem,
                   full_df, top_n_df, per_building_df, bucket_summary_df)

    report["generated"] = datetime.now().isoformat(timespec="seconds")
    report["project"] = project_name
    report["snapshot"] = snapshot_path.stem
    with open(stage_dir / "variance_report.json", "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, default=str)

    # Console summary
    print(f"  Construction tasks : {report['construction_task_count']:>6} "
          f"(leaves {report['construction_leaf_count']})")
    print(f"  Bucketed leaves    : {report['bucketed_leaf_count']:>6}")
    print(f"  Unbucketed leaves  : {report['unbucketed_leaf_count']:>6}")
    print(f"  Rolled-up lines    : {report['rolled_up_line_count']:>6}")
    print(f"  Buckets with data  : {report['buckets_with_data']:>6}")

    print("\n  Bucket net variances:")
    for _, r in bucket_summary_df.iterrows():
        print(f"    {r['bucket']:<18} net {r['net_variance']:>8.1f} d   "
              f"({r['over_count']} over +{r['over_days']:.1f} / "
              f"{r['under_count']} under {r['under_days']:.1f})")

    if report.get("site_work_net"):
        sw = report["site_work_net"]
        print(f"\n  Site Work check (Harrison ref: 14 over +452.7 / 5 under -67.8 = +384.9):")
        print(f"    got: {sw['over_count']} over +{sw['over_days']:.1f} / "
              f"{sw['under_count']} under {sw['under_days']:.1f} = +{sw['net']:.1f}")

    print(f"\n{'='*60}")
    print(f"  Workbook : {xlsx_path}")
    print(f"  Parquet  : {stage_dir}")
    print(f"{'='*60}\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
